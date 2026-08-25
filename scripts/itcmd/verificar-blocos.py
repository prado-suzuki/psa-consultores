# -*- coding: utf-8 -*-
"""Varre qualquer WP de ITCMD/MT, localiza todos os blocos de escalonamento e
confronta cada total contra a forma fechada da spec.

Uso:
    python scripts/itcmd/verificar-blocos.py "<arq1.xlsx>" ["<arq2.xlsx>" ...]

O layout do bloco e sempre o mesmo nestas pastas: uma celula com o texto
"Escalonamento" na coluna B, na linha r. A partir dela:
    C(r-1)          -> UPF
    E..J(r-1)       -> bases (100% / 70% para contabil, ITR e mercado)
    C(r+1)..C(r+5)  -> aliquotas
    D(r+1)..D(r+4)  -> tetos das faixas
    E..J(r+6)       -> totais de imposto

Requer: openpyxl
"""
import sys
from decimal import Decimal as D, ROUND_HALF_UP

import openpyxl

FAIXAS = [(D(500), D("0.00"), D(0)), (D(1000), D("0.02"), D(10)),
          (D(4000), D("0.04"), D(30)), (D(10000), D("0.06"), D(110)),
          (None, D("0.08"), D(310))]
COLS = ["E", "F", "G", "H", "I", "J"]
ROTULOS = ["contábil 100%", "contábil 70%", "ITR 100%",
           "ITR 70%", "mercado 100%", "mercado 70%"]


def imposto(base, upf):
    base, upf = D(str(base)), D(str(upf))
    for lim, aliq, ded in FAIXAS:
        if lim is None or base <= lim * upf:
            v = aliq * base - ded * upf
            return v.quantize(D("0.01"), ROUND_HALF_UP), v < 0
    raise AssertionError


def num(v):
    if v is None or isinstance(v, str):
        return None
    try:
        return D(str(v))
    except Exception:
        return None


def brl(x):
    if x is None:
        return f"{'—':>15}"
    return f"{D(str(x)).quantize(D('0.01'), ROUND_HALF_UP):>15,.2f}".replace(
        ",", "X").replace(".", ",").replace("X", ".")


def blocos(ws):
    """Localiza as linhas que contem 'Escalonamento' na coluna B."""
    for row in ws.iter_rows(min_col=2, max_col=2):
        for c in row:
            if isinstance(c.value, str) and c.value.strip().lower().startswith("escalonamento"):
                yield c.row


def analisar(path):
    print(f"\n{'#' * 118}")
    print(f"# {path}")
    print(f"{'#' * 118}")
    wbv = openpyxl.load_workbook(path, data_only=True)
    wbf = openpyxl.load_workbook(path, data_only=False)
    tot_ok = tot_div = tot_neg = 0
    upfs = {}
    literais = []

    for ws in wbv.worksheets:
        wsf = wbf[ws.title]
        rs = list(blocos(ws))
        if not rs:
            continue
        print(f"\n=== aba '{ws.title}' (state={wsf.sheet_state}) — {len(rs)} bloco(s)")
        for r in rs:
            upf = num(ws[f"C{r - 1}"].value)
            rot_upf = ws[f"B{r - 1}"].value
            titulo = ws[f"B{r - 2}"].value or ws[f"B{r - 3}"].value or ""
            if upf is None:
                print(f"  linha {r}: UPF não numérica em C{r - 1} ({ws[f'C{r-1}'].value!r}) — bloco ignorado")
                continue
            upfs.setdefault(str(upf), []).append(f"{ws.title}!C{r-1}")
            print(f"\n  bloco linha {r} | UPF {brl(upf).strip()} ({rot_upf}) | {str(titulo)[:60]}")
            print(f"    {'cenário':16}{'base':>15}{'origem da base':>22}{'planilha':>15}"
                  f"{'correto':>15}{'delta':>15}")
            for col, rot in zip(COLS, ROTULOS):
                base = num(ws[f"{col}{r - 1}"].value)
                wp = num(ws[f"{col}{r + 6}"].value)
                if base is None:
                    continue
                f_base = wsf[f"{col}{r - 1}"].value
                origem = "LITERAL" if not (isinstance(f_base, str) and f_base.startswith("=")) \
                    else str(f_base)[:20]
                got, _ = imposto(base, upf)
                if origem == "LITERAL":
                    literais.append(f"{ws.title}!{col}{r - 1} = {base}")
                if wp is None:
                    print(f"    {rot:16}{brl(base)}{origem:>22}{'sem total':>15}{brl(got)}{'':>15}")
                    continue
                delta = wp - got
                if wp < 0:
                    tag = "NEGATIVO"
                    tot_neg += 1
                elif abs(delta) <= D("0.02"):
                    tag = "OK"
                    tot_ok += 1
                else:
                    tag = "DIVERGE"
                    tot_div += 1
                print(f"    {rot:16}{brl(base)}{origem:>22}{brl(wp)}{brl(got)}"
                      f"{brl(delta) if tag != 'OK' else '':>15}  {tag if tag != 'OK' else ''}")

    print(f"\n>>> resumo: {tot_ok} OK, {tot_div} divergentes, "
          f"{tot_neg} com imposto negativo na planilha")
    print(">>> UPFs encontradas nesta pasta:")
    for v, onde in sorted(upfs.items(), key=lambda t: D(t[0])):
        print(f"      R$ {brl(v).strip():>10}  em {len(onde)} bloco(s): {', '.join(onde[:6])}"
              + (" ..." if len(onde) > 6 else ""))
    if len(upfs) > 1:
        print("    ATENÇÃO: mais de uma UPF na mesma pasta — competências misturadas")

    if literais:
        print(f"\n>>> BASES DIGITADAS COMO LITERAL ({len(literais)}) — o imposto pode estar")
        print("    correto para o número digitado e ainda assim errado, porque a base não")
        print("    acompanha o quadro de quotas. Este verificador NÃO detecta esse caso:")
        for l in literais:
            print(f"      {l}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    for p in sys.argv[1:]:
        analisar(p)


if __name__ == "__main__":
    main()
