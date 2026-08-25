# -*- coding: utf-8 -*-
"""Analise do WP de ITCMD/MT que embasa docs/itcmd-mt/SPEC-motor-itcmd-mt.md.

Uso:
    python scripts/itcmd/analisar-wp.py "<caminho do .xlsx>" [--dump saida.txt]

Executa quatro analises:
  1. dump   - toda celula com formula, valor em cache e formato, abas ocultas incluidas
  2. refs   - grafo de referencias: quais entradas e saidas sao mortas
  3. bens   - reconstrucao independente dos tres totais de avaliacao
  4. faixas - confronto da forma fechada do escalonamento contra as celulas do WP

Requer: openpyxl
"""
import re
import sys
from collections import defaultdict
from decimal import Decimal as D, ROUND_HALF_UP

import openpyxl
from openpyxl.utils import column_index_from_string as ci, get_column_letter as gl

ABA_VISIVEL = "Doação"

# --- nucleo determinístico: faixas de doacao ITCMD/MT ----------------------
# (limite superior em UPF, aliquota, parcela a deduzir em UPF)
FAIXAS = [
    (D(500), D("0.00"), D(0)),
    (D(1000), D("0.02"), D(10)),
    (D(4000), D("0.04"), D(30)),
    (D(10000), D("0.06"), D(110)),
    (None, D("0.08"), D(310)),
]


def imposto(base, upf, arredondar=True):
    """Forma fechada: aliquota(k) * base - deducao_upf(k) * upf."""
    base, upf = D(str(base)), D(str(upf))
    for lim, aliq, ded in FAIXAS:
        if lim is None or base <= lim * upf:
            v = aliq * base - ded * upf
            if v < 0:
                raise ValueError(f"imposto negativo ({v}) para base {base} e UPF {upf}")
            return v.quantize(D("0.01"), ROUND_HALF_UP) if arredondar else v
    raise AssertionError("faixa nao encontrada")


def brl(x):
    s = f"{D(str(x)).quantize(D('0.01'), ROUND_HALF_UP):>18,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


# --- 1. dump ---------------------------------------------------------------
def dump(path, destino=None):
    wbf = openpyxl.load_workbook(path, data_only=False)
    wbv = openpyxl.load_workbook(path, data_only=True)
    out = []
    p = out.append

    p("### ABAS")
    for ws in wbf.worksheets:
        p(f"  '{ws.title}' | state={ws.sheet_state} | dims={ws.dimensions}")

    for ws in wbf.worksheets:
        wsv = wbv[ws.title]
        p(f"\n{'=' * 96}\nABA: '{ws.title}'  state={ws.sheet_state}\n{'=' * 96}")
        oc_r = [str(k) for k, v in ws.row_dimensions.items() if v.hidden]
        oc_c = [k for k, v in ws.column_dimensions.items() if v.hidden]
        if oc_r:
            p("LINHAS OCULTAS: " + ",".join(oc_r))
        if oc_c:
            p("COLUNAS OCULTAS: " + ",".join(oc_c))
        if ws.merged_cells.ranges:
            p("MESCLADAS: " + ", ".join(str(r) for r in ws.merged_cells.ranges))
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                cv = wsv[c.coordinate].value
                f, cvs = str(c.value), "" if cv is None else str(cv)
                nf = "" if c.number_format == "General" else f" | fmt={c.number_format}"
                if not f.startswith("=") and f == cvs:
                    p(f"  {c.coordinate} | {f}{nf}")
                else:
                    p(f"  {c.coordinate} | {f} | => {cvs}{nf}")

    txt = "\n".join(out)
    if destino:
        with open(destino, "w", encoding="utf-8") as fh:
            fh.write(txt)
        print(f"[dump] {len(txt):,} chars -> {destino}")
    else:
        print(f"[dump] {len(txt):,} chars, {txt.count(chr(10)):,} linhas "
              f"(use --dump ARQUIVO para gravar)")
    return txt


# --- 2. grafo de referencias ----------------------------------------------
CELL = re.compile(r"\$?([A-Z]{1,2})\$?(\d{1,4})")
RANGE = re.compile(r"\$?([A-Z]{1,2})\$?(\d{1,4}):\$?([A-Z]{1,2})\$?(\d{1,4})")

GRUPOS = {
    "Base de Cálculo - % (declarada)":
        ["C76", "F76", "I76", "C105", "F105", "I105",
         "C133", "F133", "I133", "C160", "F160", "I160"],
    "fator 100% do escalonamento":
        ["E91", "G91", "I91", "E120", "G120", "I120",
         "E148", "G148", "I148", "E173", "G173", "I173"],
    "fator 70% do escalonamento (usufruto)":
        ["F91", "H91", "J91", "F120", "H120", "J120",
         "F148", "H148", "J148", "F173", "H173", "J173"],
    "UPF": ["C90", "C119", "C147", "C172"],
    "saídas do ramo de 70%":
        ["F97", "H97", "J97", "F126", "H126", "J126",
         "F154", "H154", "J154", "F179", "H179", "J179"],
    "'o que vai na apresentação'": ["L127", "M127", "N127"],
}


def refs(path):
    ws = openpyxl.load_workbook(path, data_only=False)[ABA_VISIVEL]
    usado_por = defaultdict(set)
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            expr = v
            for m in RANGE.finditer(v):
                c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
                for col in range(ci(c1), ci(c2) + 1):
                    for r in range(r1, r2 + 1):
                        usado_por[f"{gl(col)}{r}"].add(c.coordinate)
                expr = expr.replace(m.group(0), " ")
            for m in CELL.finditer(expr):
                usado_por[f"{m.group(1)}{m.group(2)}"].add(c.coordinate)

    mortas = 0
    for grupo, cells in GRUPOS.items():
        print(f"\n[refs] {grupo}")
        for c in cells:
            u = sorted(usado_por.get(c, ()))
            if not u:
                mortas += 1
            estado = "MORTA" if not u else "usada por " + ", ".join(u)
            print(f"       {c:5} = {str(ws[c].value):<22} -> {estado}")
    print(f"\n[refs] total de células declaradas e não referenciadas: {mortas}")


# --- 3. bens --------------------------------------------------------------
def bens(path):
    ws = openpyxl.load_workbook(path, data_only=True)[ABA_VISIVEL]
    tc = ti = tm = D(0)
    sem_itr, areas = [], defaultdict(list)
    print(f"\n[bens] {'linha':>5} {'matrícula':>10} {'área':>11} "
          f"{'contábil':>16} {'ITR':>16} {'mercado':>18}")
    for r in range(21, 32):
        cont, itr, merc = ws[f"F{r}"].value, ws[f"G{r}"].value, ws[f"H{r}"].value
        area, matr = ws[f"D{r}"].value, ws[f"C{r}"].value
        if cont is None:
            continue
        tc += D(str(cont))
        tm += D(str(merc or 0))
        if itr is None:
            sem_itr.append((r, matr, area))
        else:
            ti += D(str(itr))
        if area is not None:
            areas[matr].append((r, str(area)))
        fitr = "AUSENTE" if itr is None else brl(itr)
        print(f"[bens] {r:>5} {str(matr):>10} {str(area):>11} "
              f"{brl(cont)} {fitr:>16} {brl(merc or 0)}")

    moeda = sum((D(str(ws[f"J{r}"].value or 0)) for r in (28, 29, 30)), D(0))
    print(f"[bens] {'':5} {'MOEDA CORRENTE':>10} {'':11} "
          f"{brl(moeda)} {brl(moeda)} {brl(moeda)}")
    print(f"[bens] {'':5} {'RECONSTRUÍDO':>10} {'':11} "
          f"{brl(tc + moeda)} {brl(ti + moeda)} {brl(tm + moeda)}")
    print(f"[bens] {'':5} {'WP F34/G34/H34':>10} {'':11} "
          f"{brl(ws['F34'].value)} {brl(ws['G34'].value)} {brl(ws['H34'].value)}")

    confere = all(D(str(a)).quantize(D("0.01")) == D(str(b)).quantize(D("0.01"))
                  for a, b in ((tc + moeda, ws["F34"].value),
                               (ti + moeda, ws["G34"].value),
                               (tm + moeda, ws["H34"].value)))
    print(f"[bens] reconstrução confere com o WP: {'SIM' if confere else 'NÃO'}")

    def ha(s):
        return D(str(s).replace(",", "."))

    perdida = sum((ha(a) for _, _, a in sem_itr if a), D(0))
    total_ha = sum((ha(a) for occ in areas.values() for _, a in occ), D(0))
    print(f"\n[bens] LACUNAS DE ITR ({len(sem_itr)} linhas, "
          f"{float(perdida):,.4f} ha de {float(total_ha):,.4f} ha = "
          f"{float(perdida / total_ha):.1%} da área):")
    for r, matr, area in sem_itr:
        print(f"       linha {r}: matrícula {matr}, {area} ha — coluna G vazia")
    for matr, occ in areas.items():
        if len(occ) > 1:
            print(f"[bens] MATRÍCULA REPETIDA {matr}: " +
                  " e ".join(f"linha {r} ({a} ha)" for r, a in occ))

    vq_itr = D(str(ws["G34"].value)) / D(str(ws["F34"].value))
    vq_mkt = D(str(ws["H34"].value)) / D(str(ws["F34"].value))
    print(f"[bens] valor/quota exato — ITR {vq_itr}  mercado {vq_mkt}")


# --- 4. faixas ------------------------------------------------------------
# (rotulo, celula da base, celula do total) por bloco da aba visivel
BLOCOS = [
    ("consolidado",         90, 97),
    ("Cristiano",          119, 126),
    ("Fabiane",            147, 154),
    ("Cristiano integral", 172, 179),
]
COLS = [("contábil 100%", "E"), ("contábil 70%", "F"), ("ITR 100%", "G"),
        ("ITR 70%", "H"), ("mercado 100%", "I"), ("mercado 70%", "J")]


def faixas(path):
    ws = openpyxl.load_workbook(path, data_only=True)[ABA_VISIVEL]
    print(f"\n[faixas] {'bloco / cenário':34} {'base':>18} {'WP':>18} "
          f"{'forma fechada':>18} {'delta':>16}  ok")
    ok = div = 0
    for rot, rbase, rtot in BLOCOS:
        upf = ws[f"C{rbase}"].value
        for nome, col in COLS:
            base, wp = ws[f"{col}{rbase}"].value, ws[f"{col}{rtot}"].value
            if base is None or wp is None:
                continue
            try:
                got = imposto(base, upf)
                gots = brl(got)
            except ValueError:
                got, gots = None, "ERRO (negativo)"
            delta = None if got is None else D(str(wp)) - got
            confere = delta is not None and abs(delta) < D("0.005")
            ok, div = (ok + 1, div) if confere else (ok, div + 1)
            print(f"[faixas] {rot + ' · ' + nome:34} {brl(base)} {brl(wp)} "
                  f"{gots:>18} {'' if delta is None else brl(delta):>16}  "
                  f"{'OK' if confere else 'DIVERGE'}")
    print(f"\n[faixas] UPF do WP: {ws['C90'].value} "
          f"(as 4 células C90/C119/C147/C172 são literais independentes)")
    print(f"[faixas] {ok} células conferem, {div} divergem "
          f"— as divergentes estão especificadas em G2 do golden-master")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(2)
    path = sys.argv[1]
    destino = None
    if "--dump" in sys.argv:
        destino = sys.argv[sys.argv.index("--dump") + 1]
    dump(path, destino)
    refs(path)
    bens(path)
    faixas(path)


if __name__ == "__main__":
    main()
